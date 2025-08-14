import cv2
import numpy as np
import random
import time
import os
import openpyxl

def extract_subject(frame):
    """
    緑色の比率のみを使って被写体を抽出:
    G/(R+G+B) > 0.35（全体に対する緑の比率）
    """
    # チャンネルをfloat32型に変換して演算
    frame = frame.astype(np.float32)
    
    # BGRからチャンネルを分離（OpenCVはBGR形式）
    b = frame[:,:,0]
    g = frame[:,:,1]
    r = frame[:,:,2]
    
    # RGB合計（ゼロ除算回避のため小さな値を加える）
    rgb_sum = r + g + b + 1e-10
    
    # 緑の比率計算
    g_ratio = g / rgb_sum
    
    # 条件: G比率が閾値以上
    G_RATIO_THRESHOLD = 0.30
    mask = (g_ratio > G_RATIO_THRESHOLD).astype(np.uint8) * 255
    
    # デバッグ用にマスクを保存
    cv2.imwrite('debug_g_ratio_mask.png', mask)
    
    # ノイズ除去
    kernel = np.ones((5,5), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    
    # ノイズ除去後のマスク保存
    cv2.imwrite('debug_mask_after_noise.png', mask)
    
    # マスクを適用して被写体を抽出
    subject = cv2.bitwise_and(frame.astype(np.uint8), frame.astype(np.uint8), mask=mask)
    
    # 抽出結果の保存
    cv2.imwrite('debug_subject.png', subject)
    
    return subject

def findSquares(bin_image, image, cond_area = 675):
    """
    多角形検出関数
    """
    # 輪郭取得
    contours, _ = cv2.findContours(bin_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)
    poly_num = 0
    
    for i, cnt in enumerate(contours):
        # 輪郭の周囲に比例する精度で輪郭を近似する
        arclen = cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, arclen*0.01, True)
        
        # 凸性の確認 
        area = abs(cv2.contourArea(approx))
        rcnt = approx.reshape(-1,2)
        
        if 675000 > area > cond_area and approx.shape[0] < 20:
            rcnt = approx.reshape(-1,2)
            cv2.polylines(image, [rcnt], True, (0, 0, 255), thickness=8, lineType=cv2.LINE_8)
            poly_num += 1
                
        elif area < cond_area:
            blank = np.array([rcnt])
            cv2.fillConvexPoly(bin_image, blank, (0, 0, 0))
        
    return image, poly_num

def process_image(image_path):
    # 画像を読み込む
    frame = cv2.imread(image_path)

    if frame is None:
        return f"Error: Unable to read image {image_path}", None, None, 0, 0

    # 時間計測開始
    t1 = time.time()
    
    # 被写体の抽出
    frame = extract_subject(frame)

    # グレースケール変換
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # 二値化 大津の手法
    re, th = cv2.threshold(gray, 0, 255, cv2.THRESH_OTSU)

    # 白黒領域計算
    image_size = th.size
    whitePixels = np.count_nonzero(th >= re)
    blackPixels = th.size - whitePixels
    
    whiteAreaRatio = (whitePixels/image_size)*100  # [%]
    blackAreaRatio = (blackPixels/image_size)*100  # [%]

    if whiteAreaRatio >= 101:
        # 白黒反転
        img2 = cv2.bitwise_not(th)
    else:
        img2 = th    

    # ラベリング処理
    ret, markers, stats, center = cv2.connectedComponentsWithStats(img2)

    # ラベリング結果書き出し準備
    color_src = cv2.cvtColor(img2, cv2.COLOR_GRAY2BGR)
    height, width = img2.shape[:2]
    colors = []

    # スコアリングの準備
    total_score = 0
    total_area = 0
    total_valid_pixels = 0
    
    # 輝度の有効範囲を定義（31-199）
    VALID_INTENSITY_MIN = 31
    VALID_INTENSITY_MAX = 255

    # 一定面積以下の領域を無視しつつ、スコアリング
    for i in range(1, ret):
        if stats[i][4] >= 675:
            area = stats[i][4]
            total_area += area
            colors.append(np.array([random.randint(0, 255), random.randint(0, 255), random.randint(0, 255)]))
            
            # スコアリング
            x, y, w, h = stats[i][:4]
            region = gray[y:y+h, x:x+w]
            region_mask = (markers[y:y+h, x:x+w] == i)
            
            # 有効な輝度範囲内のピクセルをカウント
            valid_pixels = np.sum(
                (region >= VALID_INTENSITY_MIN) & 
                (region <= VALID_INTENSITY_MAX) & 
                region_mask
            )
            total_valid_pixels += valid_pixels
            
            region_score = valid_pixels
            total_score += region_score
        else:
            ret = ret - 1
            colors.append(np.array([0, 0, 0]))

    # ラベリング結果と色付きスコア画像を作成
    color_src = np.zeros((height, width, 3), dtype=np.uint8)
    score_image = np.zeros((height, width, 3), dtype=np.uint8)
    
    for y in range(0, height):
        for x in range(0, width):
            if markers[y, x] > 0:
                color = colors[markers[y, x] - 1]
                color_src[y, x] = color
                
                # 輝度値が有効範囲内の場合のみ表示
                pixel_intensity = gray[y, x]
                if VALID_INTENSITY_MIN <= pixel_intensity <= VALID_INTENSITY_MAX:
                    score_image[y, x] = color
            else:
                color_src[y, x] = [0, 0, 0]
                score_image[y, x] = [0, 0, 0]

    # 結果の出力
    output_image_path = f"labeled_{os.path.basename(image_path)}"
    cv2.imwrite(output_image_path, color_src)

    # スコア画像の出力
    score_image_path = f"score_{os.path.basename(image_path)}"
    cv2.imwrite(score_image_path, score_image)

    # === 多角形検出を最後に追加 ===
    # 元の画像をコピーして多角形検出用に使用
    polygon_image = cv2.imread(image_path)
    th_copy = th.copy()  # 二値画像をコピー
    
    # 多角形検出実行
    polygon_result, poly_num = findSquares(th_copy, polygon_image)
    
    # 多角形検出結果を保存
    polygon_output_path = f"polygon_{os.path.basename(image_path)}"
    cv2.imwrite(polygon_output_path, polygon_result)

    # 処理時間表示
    t2 = time.time() - t1

    # 結果をテキストとして返す
    result = f"Image: {image_path}\n"
    result += f"閾値: {re}\n"
    result += f"White Area [%]: {whiteAreaRatio:.2f}\n"
    result += f"Black Area [%]: {blackAreaRatio:.2f}\n"
    result += f"物体の数: {ret-1}\n"
    result += f"全体の面積: {total_area}\n"
    result += f"有効輝度範囲(31-199)のピクセル数: {total_valid_pixels}\n"
    result += f"Total score: {total_score}\n"
    result += f"検出された多角形数: {poly_num}\n"  # 多角形検出結果を追加
    result += f"処理時間: {t2:.2f} s\n"
    result += f"領域の数: {ret-1}\n"
    result += f"ラベリング結果の画像: {output_image_path}\n"
    result += f"スコア画像: {score_image_path}\n"
    result += f"多角形検出画像: {polygon_output_path}\n"  # 多角形検出画像パスを追加
    result += f"\n"

    return result, color_src, score_image, ret-1, poly_num

def create_excel_file():
    """
    エクセルファイルを作成し、ヘッダーを設定
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detection Results"
    
    # ヘッダーを設定
    headers = [
        "Row", "Image Name", "Labeling Objects", "Polygon Objects", 
        "Processing Time", "Success (Labeling)", "Success (Polygon)"
    ]
    
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # エクセルファイルを保存
    workbook.save("detection_results.xlsx")
    return workbook

def save_to_excel(row_num, image_name, labeling_objects, polygon_objects, processing_time):
    """
    エクセルファイルに結果を保存
    """
    try:
        # 既存のエクセルファイルを開く
        workbook = openpyxl.load_workbook("detection_results.xlsx")
    except FileNotFoundError:
        # ファイルが存在しない場合は新規作成
        workbook = create_excel_file()
    
    sheet = workbook.active
    
    # 成功判定（1個の場合を成功とする）
    labeling_success = "成功" if labeling_objects == 1 else "失敗"
    polygon_success = "成功" if polygon_objects == 1 else "失敗"
    
    # データを書き込み
    sheet.cell(row=row_num+1, column=1, value=row_num)  # Row番号
    sheet.cell(row=row_num+1, column=2, value=image_name)  # 画像名
    sheet.cell(row=row_num+1, column=3, value=labeling_objects)  # ラベリング検出数
    sheet.cell(row=row_num+1, column=4, value=polygon_objects)  # 多角形検出数
    sheet.cell(row=row_num+1, column=5, value=round(processing_time, 3))  # 処理時間
    sheet.cell(row=row_num+1, column=6, value=labeling_success)  # ラベリング成功/失敗
    sheet.cell(row=row_num+1, column=7, value=polygon_success)  # 多角形成功/失敗
    
    # ファイルを保存
    workbook.save("detection_results.xlsx")

# メイン処理
output_file = "image_processing_results.txt"
image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']

# エクセルファイルを初期化
create_excel_file()

row_counter = 1  # エクセルの行カウンター（ヘッダー除く）

print("=== 画像処理開始 ===")

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("Image Processing Results\n")
    f.write("="*50 + "\n\n")
    
    for filename in os.listdir('.'):
        if any(filename.lower().endswith(ext) for ext in image_extensions):
            print(f"Processing {filename}...")
            
            # 画像処理実行
            result, labeled_image, score_image, labeling_objects, polygon_objects = process_image(filename)
            
            # テキストファイルに結果を書き込み
            f.write(result)
            
            # 処理時間を抽出（結果文字列から）
            processing_time = 0.0
            for line in result.split('\n'):
                if '処理時間:' in line:
                    processing_time = float(line.split(':')[1].replace('s', '').strip())
                    break
            
            # エクセルファイルに結果を保存
            save_to_excel(row_counter, filename, labeling_objects, polygon_objects, processing_time)
            
            # 結果をコンソールに表示
            print(f"  - ラベリング検出数: {labeling_objects}")
            print(f"  - 多角形検出数: {polygon_objects}")
            print(f"  - 処理時間: {processing_time:.3f}s")
            
            row_counter += 1

print(f"\n=== 処理完了 ===")
print(f"処理画像数: {row_counter-1}枚")
print(f"テキスト結果: {output_file}")
print(f"エクセル結果: detection_results.xlsx")

# 最終的な統計情報をエクセルに追加
try:
    workbook = openpyxl.load_workbook("detection_results.xlsx")
    sheet = workbook.active
    
    # 統計情報を最下行に追加
    last_row = sheet.max_row
    
    # 空行を追加
    sheet.cell(row=last_row+2, column=1, value="=== 統計情報 ===")
    
    # ラベリング成功率
    labeling_success_formula = f'=COUNTIF(F2:F{last_row},"成功")/COUNT(F2:F{last_row})*100'
    sheet.cell(row=last_row+3, column=1, value="ラベリング成功率(%)")
    sheet.cell(row=last_row+3, column=2, value=labeling_success_formula)
    
    # 多角形成功率
    polygon_success_formula = f'=COUNTIF(G2:G{last_row},"成功")/COUNT(G2:G{last_row})*100'
    sheet.cell(row=last_row+4, column=1, value="多角形成功率(%)")
    sheet.cell(row=last_row+4, column=2, value=polygon_success_formula)
    
    # 平均処理時間
    avg_time_formula = f'=AVERAGE(E2:E{last_row})'
    sheet.cell(row=last_row+5, column=1, value="平均処理時間(s)")
    sheet.cell(row=last_row+5, column=2, value=avg_time_formula)
    
    workbook.save("detection_results.xlsx")
    print("統計情報もエクセルに追加しました。")
    
except Exception as e:
    print(f"統計情報の追加でエラー: {e}")

print(f"すべての処理が完了しました。")
import cv2
import numpy as np
import random
import time
import os
import openpyxl
from pathlib import Path

def extract_subject(frame, debug_dir):
    """
    緑色の比率のみを使って被写体を抽出:
    G/(R+G+B) > 0.35（全体に対する緑の比率）
    """
    # チャンネルをfloat32型に変換して演算
    frame = frame.astype(np.float32)
    
    # BGRからチャンネルを分離（OpenCVはBGR形式）
    b = frame[:,:,0]
    g = frame[:,:,1]
    r = frame[:,:,2]
    
    # RGB合計（ゼロ除算回避のため小さな値を加える）
    rgb_sum = r + g + b + 1e-10
    
    # 緑の比率計算
    g_ratio = g / rgb_sum
    
    # 条件: G比率が閾値以上
    G_RATIO_THRESHOLD = 0.35
    mask = (g_ratio > G_RATIO_THRESHOLD).astype(np.uint8) * 255
    
    # デバッグ用にマスクを保存（フォルダ内に保存）
    cv2.imwrite(str(debug_dir / 'debug_g_ratio_mask.png'), mask)
    
    # ノイズ除去
    kernel = np.ones((5,5), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    
    # ノイズ除去後のマスク保存（フォルダ内に保存）
    cv2.imwrite(str(debug_dir / 'debug_mask_after_noise.png'), mask)
    
    # マスクを適用して被写体を抽出
    subject = cv2.bitwise_and(frame.astype(np.uint8), frame.astype(np.uint8), mask=mask)
    
    # 抽出結果の保存（フォルダ内に保存）
    cv2.imwrite(str(debug_dir / 'debug_subject.png'), subject)
    
    return subject

def findSquares(bin_image, image, cond_area = 675):
    """
    多角形検出関数
    """
    # 輪郭取得
    contours, _ = cv2.findContours(bin_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)
    poly_num = 0
    
    for i, cnt in enumerate(contours):
        # 輪郭の周囲に比例する精度で輪郭を近似する
        arclen = cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, arclen*0.01, True)
        
        # 凸性の確認 
        area = abs(cv2.contourArea(approx))
        rcnt = approx.reshape(-1,2)
        
        if 675000 > area > cond_area and approx.shape[0] < 20:
            rcnt = approx.reshape(-1,2)
            cv2.polylines(image, [rcnt], True, (0, 0, 255), thickness=8, lineType=cv2.LINE_8)
            poly_num += 1
                
        elif area < cond_area:
            blank = np.array([rcnt])
            cv2.fillConvexPoly(bin_image, blank, (0, 0, 0))
        
    return image, poly_num

def process_image_in_folder(image_path, folder_path):
    """
    フォルダ内で画像処理を実行（デバッグファイルも同じフォルダに保存）
    """
    # 画像を読み込む
    frame = cv2.imread(str(image_path))

    if frame is None:
        return f"Error: Unable to read image {image_path}", None, None, 0, 0, 0

    # 時間計測開始
    t1 = time.time()
    
    # 被写体の抽出（デバッグファイルをフォルダ内に保存）
    frame = extract_subject(frame, folder_path)

    # グレースケール変換
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # 二値化 大津の手法
    re, th = cv2.threshold(gray, 0, 255, cv2.THRESH_OTSU)

    # 白黒領域計算
    image_size = th.size
    whitePixels = np.count_nonzero(th >= re)
    blackPixels = th.size - whitePixels
    
    whiteAreaRatio = (whitePixels/image_size)*100  # [%]
    blackAreaRatio = (blackPixels/image_size)*100  # [%]

    if whiteAreaRatio >= 101:
        # 白黒反転
        img2 = cv2.bitwise_not(th)
    else:
        img2 = th    

    # ラベリング処理
    ret, markers, stats, center = cv2.connectedComponentsWithStats(img2)

    # ラベリング結果書き出し準備
    color_src = cv2.cvtColor(img2, cv2.COLOR_GRAY2BGR)
    height, width = img2.shape[:2]
    colors = []

    # スコアリングの準備
    total_score = 0
    total_area = 0
    total_valid_pixels = 0
    
    # 輝度の有効範囲を定義（31-199）
    VALID_INTENSITY_MIN = 31
    VALID_INTENSITY_MAX = 199

    # 一定面積以下の領域を無視しつつ、スコアリング
    for i in range(1, ret):
        if stats[i][4] >= 675:
            area = stats[i][4]
            total_area += area
            colors.append(np.array([random.randint(0, 255), random.randint(0, 255), random.randint(0, 255)]))
            
            # スコアリング
            x, y, w, h = stats[i][:4]
            region = gray[y:y+h, x:x+w]
            region_mask = (markers[y:y+h, x:x+w] == i)
            
            # 有効な輝度範囲内のピクセルをカウント
            valid_pixels = np.sum(
                (region >= VALID_INTENSITY_MIN) & 
                (region <= VALID_INTENSITY_MAX) & 
                region_mask
            )
            total_valid_pixels += valid_pixels
            
            region_score = valid_pixels
            total_score += region_score
        else:
            ret = ret - 1
            colors.append(np.array([0, 0, 0]))

    # ラベリング結果と色付きスコア画像を作成
    color_src = np.zeros((height, width, 3), dtype=np.uint8)
    score_image = np.zeros((height, width, 3), dtype=np.uint8)
    
    for y in range(0, height):
        for x in range(0, width):
            if markers[y, x] > 0:
                color = colors[markers[y, x] - 1]
                color_src[y, x] = color
                
                # 輝度値が有効範囲内の場合のみ表示
                pixel_intensity = gray[y, x]
                if VALID_INTENSITY_MIN <= pixel_intensity <= VALID_INTENSITY_MAX:
                    score_image[y, x] = color
            else:
                color_src[y, x] = [0, 0, 0]
                score_image[y, x] = [0, 0, 0]

    # 結果の出力（同じフォルダ内に保存）
    output_image_path = folder_path / f"labeled_{image_path.name}"
    cv2.imwrite(str(output_image_path), color_src)

    # スコア画像の出力（同じフォルダ内に保存）
    score_image_path = folder_path / f"score_{image_path.name}"
    cv2.imwrite(str(score_image_path), score_image)

    # === 多角形検出を最後に追加 ===
    # 元の画像をコピーして多角形検出用に使用
    polygon_image = cv2.imread(str(image_path))
    th_copy = th.copy()  # 二値画像をコピー
    
    # 多角形検出実行
    polygon_result, poly_num = findSquares(th_copy, polygon_image)
    
    # 多角形検出結果を保存（同じフォルダ内に保存）
    polygon_output_path = folder_path / f"polygon_{image_path.name}"
    cv2.imwrite(str(polygon_output_path), polygon_result)

    # 処理時間表示
    t2 = time.time() - t1

    # 結果をテキストとして返す
    result = f"Image: {image_path}\n"
    result += f"閾値: {re}\n"
    result += f"White Area [%]: {whiteAreaRatio:.2f}\n"
    result += f"Black Area [%]: {blackAreaRatio:.2f}\n"
    result += f"物体の数: {ret-1}\n"
    result += f"全体の面積: {total_area}\n"
    result += f"有効輝度範囲(31-199)のピクセル数: {total_valid_pixels}\n"
    result += f"Total score: {total_score}\n"
    result += f"検出された多角形数: {poly_num}\n"
    result += f"処理時間: {t2:.2f} s\n"
    result += f"領域の数: {ret-1}\n"
    result += f"ラベリング結果の画像: {output_image_path}\n"
    result += f"スコア画像: {score_image_path}\n"
    result += f"多角形検出画像: {polygon_output_path}\n"
    result += f"\n"

    return result, color_src, score_image, ret-1, poly_num, t2

def process_folder(folder_path):
    """
    1つのフォルダ内の全画像を処理
    """
    image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif']
    folder_results = []
    
    print(f"\n📁 フォルダ '{folder_path.name}' を処理中...")
    
    # フォルダ内の画像ファイルを取得
    image_files = [f for f in folder_path.iterdir() 
                   if f.is_file() and f.suffix.lower() in image_extensions]
    
    if not image_files:
        print(f"  ⚠️ フォルダ内に画像ファイルが見つかりません")
        return folder_results
    
    print(f"  📸 発見された画像数: {len(image_files)}枚")
    
    # フォルダ内テキストファイルを作成
    folder_text_file = folder_path / "processing_results.txt"
    
    with open(folder_text_file, 'w', encoding='utf-8') as f:
        f.write(f"Processing Results for {folder_path.name}\n")
        f.write("="*50 + "\n\n")
        
        for i, image_file in enumerate(image_files, 1):
            print(f"  🔄 ({i}/{len(image_files)}) {image_file.name} を処理中...")
            
            # 画像処理実行
            result, labeled_image, score_image, labeling_objects, polygon_objects, processing_time = process_image_in_folder(image_file, folder_path)
            
            # フォルダ内テキストファイルに結果を書き込み
            f.write(result)
            
            # 結果を保存
            folder_results.append({
                'folder_name': folder_path.name,
                'image_name': image_file.name,
                'full_path': image_file,
                'labeling_objects': labeling_objects,
                'polygon_objects': polygon_objects,
                'processing_time': processing_time
            })
            
            print(f"    ✅ ラベリング: {labeling_objects}個, 多角形: {polygon_objects}個, 時間: {processing_time:.3f}s")
    
    print(f"  📄 フォルダ内結果ファイル: {folder_text_file}")
    return folder_results

def create_excel_file():
    """
    エクセルファイルを作成し、ヘッダーを設定
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detection Results"
    
    # ヘッダーを設定
    headers = [
        "Row", "Folder", "Image Name", "Full Path", "Labeling Objects", "Polygon Objects", 
        "Processing Time", "Success (Labeling)", "Success (Polygon)", "Timestamp"
    ]
    
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # エクセルファイルを保存
    workbook.save("detection_results.xlsx")
    return workbook

def save_to_excel(results_list):
    """
    エクセルファイルに全結果を一括保存
    """
    # エクセルファイルを新規作成
    workbook = create_excel_file()
    sheet = workbook.active
    
    for row_num, result in enumerate(results_list, 1):
        # 成功判定（1個の場合を成功とする）
        labeling_success = "成功" if result['labeling_objects'] == 1 else "失敗"
        polygon_success = "成功" if result['polygon_objects'] == 1 else "失敗"
        
        # 現在の時刻を取得
        current_time = time.strftime("%Y-%m-%d %H:%M:%S")
        
        # データを書き込み
        sheet.cell(row=row_num+1, column=1, value=row_num)  # Row番号
        sheet.cell(row=row_num+1, column=2, value=result['folder_name'])  # フォルダ名
        sheet.cell(row=row_num+1, column=3, value=result['image_name'])  # 画像名
        sheet.cell(row=row_num+1, column=4, value=str(result['full_path']))  # フルパス
        sheet.cell(row=row_num+1, column=5, value=result['labeling_objects'])  # ラベリング検出数
        sheet.cell(row=row_num+1, column=6, value=result['polygon_objects'])  # 多角形検出数
        sheet.cell(row=row_num+1, column=7, value=round(result['processing_time'], 3))  # 処理時間
        sheet.cell(row=row_num+1, column=8, value=labeling_success)  # ラベリング成功/失敗
        sheet.cell(row=row_num+1, column=9, value=polygon_success)  # 多角形成功/失敗
        sheet.cell(row=row_num+1, column=10, value=current_time)  # タイムスタンプ
    
    # ファイルを保存
    workbook.save("detection_results.xlsx")

def get_target_folders(root_dir):
    """
    対象フォルダを取得
    """
    root_path = Path(root_dir)
    target_patterns = ['front_*', 'side_*', 'back_*']
    folders = []
    
    for pattern in target_patterns:
        for folder in root_path.glob(pattern):
            if folder.is_dir():
                folders.append(folder)
    
    return sorted(folders, key=lambda x: x.name)

# メイン処理
def main():
    print("🚀 === フォルダ別画像処理開始 ===")
    
    # 現在のディレクトリを取得
    current_dir = Path('.')
    
    # 対象フォルダを取得
    target_folders = get_target_folders(current_dir)
    
    if not target_folders:
        print("❌ 処理対象のフォルダが見つかりませんでした。")
        print("front_*, side_*, back_* パターンのフォルダがあることを確認してください。")
        return
    
    print(f"📂 処理対象フォルダ数: {len(target_folders)}個")
    for folder in target_folders:
        print(f"  - {folder.name}")
    
    # 全結果を格納するリスト
    all_results = []
    
    # 統合テキストファイルを作成
    overall_text_file = "image_processing_results_summary.txt"
    
    with open(overall_text_file, 'w', encoding='utf-8') as summary_file:
        summary_file.write("Image Processing Summary Report\n")
        summary_file.write("="*50 + "\n\n")
        
        # 各フォルダを順次処理
        for folder_index, folder_path in enumerate(target_folders, 1):
            print(f"\n📁 ({folder_index}/{len(target_folders)}) フォルダ '{folder_path.name}' を処理中...")
            
            # フォルダ内処理
            folder_results = process_folder(folder_path)
            
            # 結果を統合リストに追加
            all_results.extend(folder_results)
            
            # サマリーファイルに記録
            summary_file.write(f"=== {folder_path.name} ===\n")
            summary_file.write(f"処理画像数: {len(folder_results)}枚\n")
            
            if folder_results:
                labeling_success = sum(1 for r in folder_results if r['labeling_objects'] == 1)
                polygon_success = sum(1 for r in folder_results if r['polygon_objects'] == 1)
                avg_time = sum(r['processing_time'] for r in folder_results) / len(folder_results)
                
                summary_file.write(f"ラベリング成功: {labeling_success}/{len(folder_results)}枚 ({labeling_success/len(folder_results)*100:.1f}%)\n")
                summary_file.write(f"多角形成功: {polygon_success}/{len(folder_results)}枚 ({polygon_success/len(folder_results)*100:.1f}%)\n")
                summary_file.write(f"平均処理時間: {avg_time:.3f}秒\n")
            
            summary_file.write("\n")
    
    print(f"\n📊 === 全体処理完了 ===")
    print(f"処理フォルダ数: {len(target_folders)}個")
    print(f"処理画像総数: {len(all_results)}枚")
    
    if all_results:
        # エクセルファイルに統合結果を保存
        print("💾 エクセルファイルに統合結果を保存中...")
        save_to_excel(all_results)
        
        # 統計情報をエクセルに追加
        try:
            workbook = openpyxl.load_workbook("detection_results.xlsx")
            sheet = workbook.active
            
            last_row = sheet.max_row
            
            # 空行を追加
            sheet.cell(row=last_row+2, column=1, value="=== 統計情報 ===")
            
            # ラベリング成功率
            labeling_success_formula = f'=COUNTIF(H2:H{last_row},"成功")/COUNT(H2:H{last_row})*100'
            sheet.cell(row=last_row+3, column=1, value="ラベリング成功率(%)")
            sheet.cell(row=last_row+3, column=2, value=labeling_success_formula)
            
            # 多角形成功率
            polygon_success_formula = f'=COUNTIF(I2:I{last_row},"成功")/COUNT(I2:I{last_row})*100'
            sheet.cell(row=last_row+4, column=1, value="多角形成功率(%)")
            sheet.cell(row=last_row+4, column=2, value=polygon_success_formula)
            
            # 平均処理時間
            avg_time_formula = f'=AVERAGE(G2:G{last_row})'
            sheet.cell(row=last_row+5, column=1, value="平均処理時間(s)")
            sheet.cell(row=last_row+5, column=2, value=avg_time_formula)
            
            workbook.save("detection_results.xlsx")
            print("✅ 統計情報もエクセルに追加しました。")
            
        except Exception as e:
            print(f"⚠️ 統計情報の追加でエラー: {e}")
        
        print(f"\n📄 出力ファイル:")
        print(f"  - 統合エクセルファイル: detection_results.xlsx")
        print(f"  - 統合サマリー: {overall_text_file}")
        print(f"  - 各フォルダ内: processing_results.txt")
        print(f"  - 各フォルダ内: デバッグ画像、処理結果画像")
    
    print("\n🎉 すべての処理が完了しました！")

if __name__ == "__main__":
    main()